import json
import logging
import time
import fitz  # PyMuPDF
import requests
from celery import shared_task, chain
import os
from django.conf import settings

from django.db import models
from django.utils import timezone
from .models import (
    PEProject, 
    PEProjectDocument, 
    ExtractedFinancials, 
    QoEReport,
    CommercialAnalysis,
    OperationalAnalysis,
    AICallLog,
    RedFlagPattern,
    RedFlagFinding,
    ScoringRun,
    CriterionScore,
    ComplianceGate,
    RegulatoryChecklist,
    DealMemo,
    Fund,
    LPFundCommitment,
    LPProfile,
    FundDocument,
    ValuationModel,
    TermSheet,
    SPADraft,
    ImmutableAuditEvent,
)
from .pdf_utils import generate_capital_account_pdf, upload_pdf_to_b2
from .mail_utils import send_statement_notification
from .signals import _log_audit_event



import re


from .ai_client import AIModelClient
from .b2_utils import generate_presigned_download_url

logger = logging.getLogger(__name__)

def extract_text_from_pdf(pdf_content):
    """Extracts text from PDF bytes using PyMuPDF."""
    text = ""
    try:
        with fitz.open(stream=pdf_content, filetype="pdf") as doc:
            for page in doc:
                text += page.get_text()
    except Exception as e:
        logger.error(f"PyMuPDF error: {e}")
    return text

def extract_text_from_excel(excel_content):
    """Extracts text from Excel bytes using openpyxl."""
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed, skipping Excel extraction")
        return ""
    
    text = ""
    try:
        wb = load_workbook(io.BytesIO(excel_content), data_only=True)
        total_chars = 0
        MAX_CHARS = 500000 # 500k char safety limit
        
        for sheet in wb.worksheets:
            if total_chars > MAX_CHARS:
                break
                
            text += f"\nSheet: {sheet.title}\n"
            # Limit to first 1000 rows per sheet to avoid huge calculation sheets
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i > 1000:
                    text += "... (Large sheet truncated)\n"
                    break
                
                # Filter out None values and clean up row
                row_data = [str(cell).strip() for cell in row if cell is not None]
                if not any(row_data): continue # Skip empty rows
                
                row_text = " | ".join(row_data)
                text += row_text + "\n"
                total_chars += len(row_text)
                    
                if total_chars > MAX_CHARS:
                    text += "\n... (Total document truncated for AI safety)\n"
                    break
    except Exception as e:
        logger.error(f"Excel extraction error: {e}")
    return text

@shared_task
def extract_financials_from_document(document_id):
    """
    Task to extract 3+ years of financial data from a PDF document.
    """
    try:
        doc = PEProjectDocument.objects.get(id=document_id)
        project = doc.project
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Extraction'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        logger.info(f"Starting financial extraction for document: {doc.filename}")
        
        # 1. Get document content
        file_content = None
        if doc.local_file:
            try:
                # In Docker, local_file points to /app/media/... which is volume mounted
                with doc.local_file.open('rb') as f:
                    file_content = f.read()
                logger.info(f"Read local file for {doc.filename}")
            except Exception as e:
                logger.warning(f"Could not read local file {doc.local_file}: {e}")
        
        if not file_content:
            logger.info(f"Downloading from B2: {doc.file_key}")
            download_url = generate_presigned_download_url(doc.file_key)
            response = requests.get(download_url)
            if response.status_code == 200:
                file_content = response.content
            else:
                raise Exception(f"Failed to download from B2. Status: {response.status_code}")
            
        # 2. Extract text based on file type
        file_ext = doc.filename.split('.')[-1].lower()
        if file_ext in ['xlsx', 'xlsm', 'xltx', 'xltm']:
            doc_text = extract_text_from_excel(file_content)
        else:
            doc_text = extract_text_from_pdf(file_content)
            
        if not doc_text:
            # Fallback for scanned PDFs or empty files
            logger.warning(f"No text extracted from {doc.filename}. AI may struggle.")
        
        # 3. Call AI Client
        client = AIModelClient()
        context_data = {
            "document_text": doc_text,
            "project_name": project.legal_name
        }
        
        raw_json = client.execute_task("financial_extraction", context_data, project=project, document=doc)
        
        # 4. Parse and Save
        # Expecting JSON format: {"fiscal_years": [{"fiscal_year_bs": "2080/81", "revenue": 100, ...}], "confidence": 0.9}
        try:
            # Clean JSON if AI wrapped it in markdown code blocks
            if "```json" in raw_json:
                raw_json = raw_json.split("```json")[1].split("```")[0].strip()
            elif "```" in raw_json:
                raw_json = raw_json.split("```")[1].split("```")[0].strip()
                
            data = json.loads(raw_json)
        except Exception as e:
            logger.error(f"Failed to parse AI JSON output: {raw_json}")
            raise Exception(f"AI output parsing error: {str(e)}")

        fiscal_years = data.get("fiscal_years", [])
        if not fiscal_years:
            # Try to handle direct list if AI didn't nest it
            if isinstance(data, list):
                fiscal_years = data
            else:
                raise Exception("No fiscal years found in AI output")

        for entry in fiscal_years:
            ExtractedFinancials.objects.update_or_create(
                project=project,
                fiscal_year_bs=entry.get("fiscal_year_bs"),
                defaults={
                    "source_document": doc,
                    "revenue_npr": entry.get("revenue", 0),
                    "ebitda_npr": entry.get("ebitda", 0),
                    "net_profit_npr": entry.get("net_profit", entry.get("pat", 0)),
                    "total_assets_npr": entry.get("total_assets", 0),
                    "total_debt_npr": entry.get("total_debt", 0),
                    # Backwards compatibility
                    "revenue": entry.get("revenue", 0),
                    "ebitda": entry.get("ebitda", 0),
                    "pat": entry.get("net_profit", entry.get("pat", 0)),
                    "extraction_confidence": data.get("confidence", 0.85),
                    "raw_ai_output": entry
                }
            )
            
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Extraction'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Successfully extracted {len(fiscal_years)} years of financials for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Error in extract_financials_from_document: {str(e)}")
        # Mark as failed in progress
        try:
            progress = project.analysis_progress or {}
            progress['Extraction'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e

@shared_task
def run_qoe_analysis(project_id):
    """
    Task to run Quality of Earnings analysis using DeepSeek R1.
    """
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['QoE'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        financials = project.financials.all().order_by('fiscal_year_bs')
        
        if not financials.exists():
            raise Exception("No extracted financials found for QoE analysis")
            
        # 1. Prepare data for AI
        financial_summary = "Financial Data History:\n"
        for f in financials:
            financial_summary += (
                f"FY {f.fiscal_year_bs}: Revenue={f.revenue_npr}, EBITDA={f.ebitda_npr}, "
                f"PAT={f.net_profit_npr}, Assets={f.total_assets_npr}, Debt={f.total_debt_npr}\n"
            )
            
        # 2. Call AI Client (Routed to DeepSeek R1)
        client = AIModelClient()
        context_data = {
            "financial_data": financial_summary,
            "project_name": project.legal_name
        }
        
        report_text = client.execute_task("qoe_analysis", context_data, project=project)
        
        # 3. Determine Status (Red/Yellow/Green)
        # We can ask the AI for a structured status or parse it.
        # For now, let's use a robust keyword search or assume AI includes it.
        status = 'CLEAN'
        lower_report = report_text.lower()
        if any(word in lower_report for word in ['high risk', 'severe', 'red flag']):
            status = 'HIGH_RISK'
        elif any(word in lower_report for word in ['caution', 'yellow flag', 'moderate risk', 'adjustment needed']):
            status = 'CAUTION'
            
        # 4. Create Report
        QoEReport.objects.create(
            project=project,
            report_text=report_text,
            status=status,
            analysis_data={
                "generated_at": str(timezone.now()),
                "fiscal_years_analyzed": [f.fiscal_year_bs for f in financials]
            }
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['QoE'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"QoE Report generated for {project.legal_name} (Status: {status})"
        
    except Exception as e:
        logger.error(f"Error in run_qoe_analysis: {str(e)}")
        raise e


@shared_task
def run_commercial_analysis(project_id):
    """
    Analyzes customer concentration and market positioning.
    """
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Prepare context
        client = AIModelClient()
        
        # Aggregate multiple form responses for a richer context
        all_responses = project.form_responses.all()
        business_desc = project.legal_name
        market_info = ""
        competition = ""
        usp = ""
        
        for r in all_responses:
            step = r.step_name.lower()
            data_str = str(r.response_data)
            if 'info' in step or 'overview' in step: business_desc = data_str
            if 'market' in step: market_info = data_str
            if 'competi' in step: competition = data_str
            if 'usp' in step or 'advantage' in step: usp = data_str

        # --- NEW: SCAN PITCH DECK/BUSINESS PLAN ---
        # Expanded to include COMMERCIAL category and filename heuristics
        pitch_deck = project.documents.filter(
            models.Q(category__in=['PITCH_DECK', 'BUSINESS_PLAN', 'COMMERCIAL']) |
            models.Q(filename__icontains='deck') |
            models.Q(filename__icontains='pitch')
        ).first()
        
        pitch_deck_content = ""
        if pitch_deck:
            logger.info(f"--- ATTEMPTING PITCH DECK SCAN: {pitch_deck.filename} ---")
            try:
                # We reuse the AI client to summarize the pitch deck specifically for commercial strategy
                pitch_deck_content = client.execute_task(
                    "document_qualitative_summary", 
                    {"filename": pitch_deck.filename}, 
                    project=project,
                    document=pitch_deck
                )
                logger.info(f"--- PITCH DECK SCAN SUCCESS: {len(pitch_deck_content)} chars ---")
            except Exception as scan_err:
                logger.error(f"--- PITCH DECK SCAN FAILED: {str(scan_err)} ---")
                pitch_deck_content = "Document scanning failed. Relying on form responses."
        else:
            logger.warning("--- NO PITCH DECK FOUND FOR SCANNING ---")

        context_data = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "business_desc": business_desc,
            "market_info": market_info,
            "competition": competition,
            "usp": usp,
            "pitch_deck_insights": pitch_deck_content,
            "market_data": f"Sector: {project.get_sector_display()}. Analysis requested for Nepal market context. {market_info} {pitch_deck_content}",
        }

        # Call AI (Routed to Gemini)
        raw_json = client.execute_task("commercial_analysis", context_data, project=project)
        
        try:
            if "```json" in raw_json:
                raw_json = raw_json.split("```json")[1].split("```")[0].strip()
            data = json.loads(raw_json)
        except:
            data = {"customer_concentration_pct": 0, "top_customer_names": "", "market_positioning_notes": raw_json}

        # Create Record
        CommercialAnalysis.objects.create(
            project=project,
            customer_concentration_pct=data.get("customer_concentration_pct", 0),
            top_customer_names=data.get("top_customer_names", ""),
            market_positioning_notes=data.get("market_positioning_notes", ""),
            ai_call_log=AICallLog.objects.filter(project=project, task_type='commercial_analysis').first()
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Commercial'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Commercial Analysis completed for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Error in run_commercial_analysis: {str(e)}")
        raise e


@shared_task
def run_operational_analysis(project_id, manual_context=""):
    """
    Analyzes technology stack and operational risks.
    """
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Operational'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        client = AIModelClient()
        
        # 1. Aggregate Operational Context from Form Responses
        tech_info = ""
        team_info = ""
        ops_info = ""
        
        responses = project.form_responses.all()
        for resp in responses:
            step = resp.step_name.lower()
            data_str = json.dumps(resp.response_data)
            if 'tech' in step or 'product' in step: tech_info = data_str
            if 'team' in step or 'founder' in step: team_info = data_str
            if 'operation' in step or 'supply' in step: ops_info = data_str

        # --- NEW: SCAN DOCUMENTS FOR OPERATIONAL DATA ---
        op_doc = project.documents.filter(
            models.Q(category__in=['PITCH_DECK', 'BUSINESS_PLAN', 'TECHNICAL', 'COMMERCIAL', 'OPERATIONAL_AUDIT', 'TECH_STACK', 'ORG_CHART']) |
            models.Q(filename__icontains='tech') |
            models.Q(filename__icontains='ops')
        ).first()
        
        doc_insights = ""
        if op_doc:
            try:
                doc_insights = client.execute_task(
                    "document_qualitative_summary", 
                    {"filename": op_doc.filename, "focus": "technology stack, operational logic, and founder expertise"}, 
                    project=project,
                    document=op_doc
                )
            except:
                doc_insights = "No document insights available."

        context_data = {
            "project_name": project.legal_name,
            "tech_stack_raw": tech_info,
            "team_details": team_info,
            "operational_notes": ops_info,
            "manual_context": manual_context,
            "document_insights": doc_insights,
            "market_sector": project.get_sector_display()
        }

        # 2. Call AI (Generates JSON + Markdown)
        # We append a strict formatting instruction to ensure the AI always provides the JSON block
        context_data["formatting_instruction"] = """
        IMPORTANT: Your response MUST end with a JSON block enclosed in ```json ... ``` tags.
        The JSON must follow this exact structure:
        {
          "technology_stack": { "Frontend": "...", "Backend": "...", "Database": "...", "Infrastructure": "..." },
          "key_person_risk_score": 1-10 integer,
          "supply_chain_risks": ["Risk 1", "Risk 2"],
          "operational_red_flags": ["Blocker 1"]
        }
        Do not include any text after the JSON block.
        """
        
        raw_output = client.execute_task("operational_analysis", context_data, project=project)
        
        data = {
            "technology_stack": {}, 
            "key_person_risk_score": 5, 
            "supply_chain_risks": [], 
            "operational_red_flags": [],
            "thesis_markdown": raw_output
        }

        try:
            # 1. Try standard markdown code block extraction
            json_str = None
            if "```json" in raw_output:
                json_str = raw_output.split("```json")[1].split("```")[0].strip()
            elif "{" in raw_output and "}" in raw_output:
                # Fallback: Find the last JSON-like structure in the text
                import re
                matches = re.findall(r'\{.*\}', raw_output, re.DOTALL)
                if matches:
                    json_str = matches[-1]

            if json_str:
                parsed = json.loads(json_str)
                data.update(parsed)
                # Aggressively clean up markdown: remove everything from the first ```json to the end
                if "```json" in raw_output:
                    data["thesis_markdown"] = raw_output.split("```json")[0].strip()
                elif json_str in raw_output:
                    data["thesis_markdown"] = raw_output.split(json_str)[0].strip()
            else:
                data["operational_red_flags"] = ["AI provided insights but skipped structured metrics. Please review thesis for details."]
        except Exception as e:
            logger.warning(f"Operational JSON parsing failed: {e}")

        # Create Record
        OperationalAnalysis.objects.create(
            project=project,
            technology_stack=data.get("technology_stack", {}),
            key_person_risk_score=data.get("key_person_risk_score", 5),
            supply_chain_risks=data.get("supply_chain_risks", []),
            operational_red_flags=data.get("operational_red_flags", []),
            thesis_markdown=data.get("thesis_markdown", raw_output),
            ai_call_log=AICallLog.objects.filter(project=project, task_type='operational_analysis').first()
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Operational'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Operational Analysis completed for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Error in run_operational_analysis: {str(e)}")
        raise e

@shared_task
def scan_legal_document(document_id):
    """
    Scans a legal document for red flags using regex patterns + Gemini Flash analysis.
    """
    try:
        doc_obj = PEProjectDocument.objects.get(id=document_id)
        project = doc_obj.project
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Legal Scan'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        # 1. Download and Extract Text
        url = generate_presigned_download_url(doc_obj.file_key)
        resp = requests.get(url)
        if resp.status_code != 200:
            raise Exception(f"Failed to download document: {resp.status_code}")
            
        text = extract_text_from_pdf(resp.content)
        if not text:
            return "No text extracted from document."

        # 2. Regex Matching
        patterns = RedFlagPattern.objects.all()
        findings_to_create = []
        
        client = AIModelClient()
        
        for p in patterns:
            matches = list(re.finditer(p.pattern_regex, text, re.IGNORECASE))
            if matches:
                # Take first match context for AI analysis
                match = matches[0]
                start = max(0, match.start() - 200)
                end = min(len(text), match.end() + 200)
                snippet = text[start:end]
                
                # Use Gemini to analyze the specific snippet
                ai_prompt = {
                    "snippet": snippet,
                    "pattern_name": p.name,
                    "description": p.description,
                    "nepal_context": p.nepal_context_note
                }
                
                analysis = client.execute_task("legal_scan", ai_prompt, project=project)
                
                findings_to_create.append(RedFlagFinding(
                    project=project,
                    document=doc_obj,
                    pattern=p,
                    context_snippet=snippet,
                    severity=p.severity,
                    ai_analysis=analysis
                ))

        if findings_to_create:
            RedFlagFinding.objects.bulk_create(findings_to_create)
            
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Legal Scan'] = 'completed'
        project.analysis_progress = progress
        project.save()
            
        return f"Legal scan completed for {doc_obj.filename}. Found {len(findings_to_create)} red flags."

    except Exception as e:
        logger.error(f"Error in scan_legal_document: {str(e)}")
        raise e
        raise e


@shared_task
def run_finlo_scoring(project_id, user_id=None):
    """
    Executes the 20-criteria FINLO scoring framework.
    Runs 5 parallel AI calls (one per pillar).
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Scoring'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        user = User.objects.get(id=user_id) if user_id else None
        
        # 1. Create Scoring Run
        run = ScoringRun.objects.create(
            project=project,
            triggered_by=user,
            status='PROCESSING'
        )
        
        # 2. Gather Context
        financials = list(project.financials.all())
        qoe = project.qoe_reports.first()
        red_flags = list(project.red_flags.all())
        responses = list(project.form_responses.all())
        operational = project.operational_analyses.first()
        commercial = project.commercial_analyses.first()
        
        context = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "deal_type": project.get_deal_type_display(),
            "financials": [
                {
                    "year": f.fiscal_year_bs,
                    "rev": float(f.revenue),
                    "ebitda": float(f.ebitda),
                    "pat": float(f.pat)
                } for f in financials
            ],
            "qoe_status": qoe.status if qoe else "N/A",
            "legal_red_flags": [f"{f.pattern.name}: {f.severity}" for f in red_flags if f.pattern],
            "operational_thesis": operational.thesis_markdown if operational else "No operational audit performed yet.",
            "operational_red_flags": operational.operational_red_flags if operational else [],
            "commercial_notes": commercial.market_positioning_notes if commercial else "No commercial audit performed yet.",
            "submission_data": [str(r.response_data) for r in responses]
        }
        
        # 3. Define Pillars
        pillars = {
            "F": {
                "name": "Foresight",
                "criteria": ["problem_clarity", "market_opportunity", "innovation_type", "ai_native_assessment"]
            },
            "I": {
                "name": "Insight",
                "criteria": ["market_size_accuracy", "revenue_model_clarity", "unit_economics_awareness", "capital_efficiency"]
            },
            "N": {
                "name": "Nexus",
                "criteria": ["founder_market_fit", "team_completeness", "leadership_conviction", "network_strength"]
            },
            "L": {
                "name": "Logic",
                "criteria": ["validation_evidence", "traction_quality", "risk_awareness", "comparable_analysis"]
            },
            "O": {
                "name": "Odyssey",
                "criteria": ["partnership_quality", "competitive_moat", "growth_vision", "finlogic_alignment"]
            }
        }
        
        client = AIModelClient()
        pillar_averages = {}
        
        # 4. Execute Consolidated AI Call
        # We consolidate all 5 pillars into 1 call for 70% token savings and holistic reasoning.
        prompt_data = {
            "pillars_definition": pillars,
            "project_data": context,
            "precision_guardrails": (
                "1. Analyze all 20 criteria across the 5 pillars (F, I, N, L, O).\n"
                "2. Every score MUST be an integer between 1-10.\n"
                "3. Provide specific evidence quotes for EACH criterion.\n"
                "4. Ensure cross-pillar consistency: if a Red Flag exists in Legal, it must reflect in Risk Awareness scoring.\n"
                "5. BREVITY: Keep rationales to max 2 concise, high-impact sentences. Avoid filler text."
            )
        }
        
        # This uses the 'scoring' task type in PromptLibrary
        raw_json = client.execute_task("scoring", prompt_data, project=project)
        
        try:
            # Aggressive markdown cleanup
            clean_json = raw_json
            if "```json" in clean_json:
                clean_json = clean_json.split("```json")[1].split("```")[0].strip()
            elif "```" in clean_json:
                clean_json = clean_json.split("```")[1].split("```")[0].strip()
            
            data = json.loads(clean_json)
            pillar_averages = {}
            
            # The AI might return the results nested under 'pillar_results' or at the root
            pillar_results = data.get("pillar_results", data)
            
            for code, info in pillars.items():
                # Try to find pillar data by code (F, I, N, L, O) or by name (Foresight, etc.)
                pillar_data = pillar_results.get(code, pillar_results.get(info["name"], {}))
                
                # If the AI returned criteria_scores at the root of the pillar object
                criteria_data = pillar_data.get("criteria_scores", pillar_data)
                
                scores_sum = 0
                count = 0
                
                for crit_id in info["criteria"]:
                    result = criteria_data.get(crit_id, {})
                    score_val = result.get("score", 5)
                    
                    CriterionScore.objects.create(
                        scoring_run=run,
                        pillar=code,
                        criterion_id=crit_id,
                        ai_score=score_val,
                        ai_rationale=result.get("rationale", "No rationale provided."),
                        ai_confidence=result.get("confidence", 0.7),
                        evidence_quotes=result.get("evidence", [])
                    )
                    scores_sum += score_val
                    count += 1
                
                pillar_averages[code] = (scores_sum / count) if count > 0 else 5

        except Exception as e:
            logger.error(f"Failed to parse consolidated scoring response: {str(e)}")
            
            # CRITICAL DEBUG: Write the failing response to a file so we can analyze it
            try:
                debug_path = os.path.join(settings.BASE_DIR, "scratch", "last_failed_scoring.json")
                os.makedirs(os.path.dirname(debug_path), exist_ok=True)
                with open(debug_path, "w", encoding="utf-8") as f:
                    f.write(raw_json)
                logger.info(f"Failing AI response dumped to: {debug_path}")
            except:
                pass

            # Create default entries so the UI isn't empty
            for code, info in pillars.items():
                for crit_id in info["criteria"]:
                    CriterionScore.objects.create(
                        scoring_run=run,
                        pillar=code,
                        criterion_id=crit_id,
                        ai_score=5,
                        ai_rationale="Consolidated scoring failed to parse. Using baseline defaults.",
                        ai_confidence=0.1
                    )
                pillar_averages[code] = 5

        # 5. Calculate Weighted Final Score
        # F: 20%, I: 25%, N: 20%, L: 20%, O: 15%
        final_score = (
            pillar_averages.get("F", 5) * 0.20 +
            pillar_averages.get("I", 5) * 0.25 +
            pillar_averages.get("N", 5) * 0.20 +
            pillar_averages.get("L", 5) * 0.20 +
            pillar_averages.get("O", 5) * 0.15
        )
        
        run.total_deal_score = round(final_score, 2)
        run.status = 'COMPLETED'
        run.save()
        
        # 6. Initialize Compliance Gates
        gates = ['FITTA', 'AML_KYC', 'FINANCIAL_AUDIT', 'LEGAL_STRUCTURE', 'SEBON_MAPPING']
        for g in gates:
            ComplianceGate.objects.get_or_create(scoring_run=run, gate_id=g)
            
        # Note: Scoring no longer changes project status automatically.
        # GP advances to IC_REVIEW manually after reviewing scoring + clearing compliance gates.
        project.save()
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Scoring'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"FINLO Scoring completed for {project.legal_name}. Score: {run.total_deal_score}"
        
    except Exception as e:
        logger.error(f"Error in run_finlo_scoring: {str(e)}")
        if 'run' in locals():
            run.status = 'FAILED'
            run.save()
        raise e


@shared_task
def run_nepal_compliance_check(project_id):
    """
    Analyzes project context to pre-populate the Nepal Regulatory Checklist.
    Detects FITTA, NRB, and SEBON requirements.
    """
    project = PEProject.objects.get(id=project_id)
    
    # 1. Detect Foreign Investment (FITTA)
    # Checks: "foreign" in company name, or if there's a specific form response
    is_foreign = False
    if "foreign" in project.legal_name.lower():
        is_foreign = True
        
    # Check form responses for keywords
    for resp in project.form_responses.all():
        data_str = json.dumps(resp.response_data).lower()
        if any(kw in data_str for kw in ["fdi", "foreign", "non-nepali", "international investor"]):
            is_foreign = True
            break
            
    # 2. Detect NRB Approval (usually for financial/banking/large scale FDI)
    is_nrb = is_foreign # Most FDI needs NRB recording/approval
    
    # 3. Create or Update Checklist
    checklist, created = RegulatoryChecklist.objects.get_or_create(
        project=project,
        defaults={
            "fitta_approval_required": is_foreign,
            "nrb_approval_required": is_nrb,
            "sebon_reporting_compliant": True, # Assume compliant initially
            "industry_specific_license_required": project.sector in [PEProject.Sector.MANUFACTURING, PEProject.Sector.AGRICULTURE]
        }
    )
    
    if not created:
        # Update if not manually locked (could add a lock flag later)
        checklist.fitta_approval_required = is_foreign
        checklist.nrb_approval_required = is_nrb
        checklist.save()
        
    return f"Compliance checklist pre-populated for {project.legal_name}"


@shared_task
def generate_memo_draft(project_id):
    """
    Gathers all project data and uses DeepSeek R1 to draft a professional investment memo.
    """
    project = PEProject.objects.get(id=project_id)
    
    # Update progress
    progress = project.analysis_progress or {}
    progress['Memo'] = 'processing'
    project.analysis_progress = progress
    project.save()
    
    # 1. Gather Context
    financials = list(project.financials.values())
    scoring = project.scoring_runs.first()
    red_flags = list(project.red_flags.values('pattern__name', 'severity', 'ai_analysis'))
    commercial = project.commercial_analyses.first()
    operational = project.operational_analyses.first()
    regulatory = getattr(project, 'regulatory_checklist', None)
    
    # Get latest valuation models
    latest_dcf = project.valuations.filter(model_type='DCF').order_by('-created_at').first()
    latest_lbo = project.valuations.filter(model_type='LBO').order_by('-created_at').first()
    
    context_data = {
        "project_name": project.legal_name,
        "sector": project.sector,
        "deal_type": project.deal_type,
        "financials": financials,
        "dcf": latest_dcf.outputs if latest_dcf else {},
        "lbo": latest_lbo.outputs if latest_lbo else {},
        "lbo_assumptions": latest_lbo.assumptions if latest_lbo else {},
        "scoring_summary": {
            "total_score": scoring.total_deal_score if scoring else "N/A",
            "pillars": list(scoring.criteria_scores.values('pillar', 'criterion_id', 'gp_score', 'ai_score')) if scoring else []
        },
        "red_flags": red_flags,
        "commercial_analysis": {
            "top_customers": commercial.top_customer_names if commercial else [],
            "market_notes": commercial.market_positioning_notes if commercial else ""
        } if commercial else None,
        "operational_analysis": {
            "risks": operational.operational_red_flags if operational else [],
            "tech_stack": operational.technology_stack if operational else {}
        } if operational else None,
        "regulatory": {
            "fitta": regulatory.fitta_approval_required if regulatory else False,
            "nrb": regulatory.nrb_approval_required if regulatory else False
        } if regulatory else None
    }
    
    # 2. Call AI Client (Routing to DeepSeek R1 for memo drafting)
    client = AIModelClient()
    raw_json = client.execute_task("memo_draft", context_data, project=project)
    
    # 3. Parse and Save
    try:
        # Clean JSON if AI wrapped it in markdown
        if "```json" in raw_json:
            raw_json = raw_json.split("```json")[1].split("```")[0].strip()
        elif "```" in raw_json:
            raw_json = raw_json.split("```")[1].split("```")[0].strip()
            
        data = json.loads(raw_json)
        
        # Ensure we have a version number
        latest_memo = project.memos.order_by('-version').first()
        new_version = (latest_memo.version + 1) if latest_memo else 1
        
        memo = DealMemo.objects.create(
            project=project,
            content=data,
            status='DRAFT',
            version=new_version
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Memo'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Memo v{new_version} drafted for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Failed to parse memo draft JSON: {str(e)}")
        raise e


@shared_task
def run_full_analysis(project_id, user_id=None):
    """
    Triggers the complete AI due diligence pipeline in sequence.
    """
    try:
        from .models import PEProjectDocument
        project = PEProject.objects.get(id=project_id)
        
        # Initialize Progress
        project.analysis_progress = {
            "Extraction": "pending",
            "QoE": "pending",
            "Commercial": "pending",
            "Operational": "pending",
            "Compliance": "pending",
            "Legal Scan": "pending",
            "Scoring": "pending",
            "Memo": "pending"
        }
        project.save()
        
        # 1. Find the most recent financial document
        doc = project.documents.filter(category='FINANCIAL').order_by('-uploaded_at').first()
        legal_doc = project.documents.filter(category='LEGAL').order_by('-uploaded_at').first()
        
        if not doc:
            return f"Full analysis aborted: No financial document found for {project.legal_name}"

        # 2. Build the chain dynamically
        steps = [
            extract_financials_from_document.si(doc.id),
            run_qoe_analysis.si(project.id),
            run_commercial_analysis.si(project.id),
            run_operational_analysis.si(project.id),
            run_nepal_compliance_check.si(project.id)
        ]
        
        if legal_doc:
            steps.append(scan_legal_document.si(legal_doc.id))
            
        steps.append(run_finlo_scoring.si(project.id))
        steps.append(generate_memo_draft.si(project.id))
        
        analysis_chain = chain(*steps)
        
        analysis_chain.apply_async()
        
        # Log Audit Event
        from .signals import _log_audit_event
        from django.contrib.auth import get_user_model
        User = get_user_model()
        actor = User.objects.filter(id=user_id).first() if user_id else None

        _log_audit_event(
            event_type='FULL_ANALYSIS_TRIGGERED',
            actor=actor,
            obj=project,
            payload={"document_id": str(doc.id) if doc else None}
        )
        
        return f"Full analysis chain started for {project.legal_name}"
    except Exception as e:
        logger.error(f"Error in run_full_analysis: {str(e)}")
        # Mark all steps as failed
        try:
            from .models import PEProject
            project = PEProject.objects.get(id=project_id)
            progress = project.analysis_progress or {}
            for k in progress:
                if progress[k] in ['processing', 'pending']:
                    progress[k] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e


@shared_task
def generate_lp_statements(fund_id, quarter, year, lpprofile_id=None, gp_user_id=None):
    """
    Generate capital account statements for one or all LPs in a fund.
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    try:
        fund = Fund.objects.get(id=fund_id)
        gp_user = User.objects.get(id=gp_user_id) if gp_user_id else None
        
        # Determine LPs to process
        if lpprofile_id:
            commitments = LPFundCommitment.objects.filter(fund=fund, lp_profile_id=lpprofile_id)
        else:
            commitments = LPFundCommitment.objects.filter(fund=fund)
            
        count = 0
        for lp_commitment in commitments:
            lp_profile = lp_commitment.lp_profile
            file_name = f"Statement_{fund.name.replace(' ', '_')}_{quarter}_{year}_{lp_profile.id}"
            
            # 1. Generate PDF
            pdf_bytes = generate_capital_account_pdf(lp_commitment, quarter, year)
            
            # 2. Upload to B2 and create record
            upload_pdf_to_b2(pdf_bytes, file_name, fund, lp_commitment, quarter, year, gp_user)
            
            # 3. Send email notification
            send_statement_notification(lp_profile, fund, quarter, year)
            
            count += 1
            
        return {
            "status": "success",
            "message": f"Generated {count} statements for {fund.name} ({quarter} {year})",
            "count": count
        }
    except Exception as e:
        logger.error(f"Error in generate_lp_statements: {e}")
        return {
            "status": "error",
            "message": str(e)
        }




@shared_task
def move_project_documents_to_b2(project_id):
    """
    Moves all locally stored documents for a project to Backblaze B2.
    """
    from .b2_utils import upload_local_file_to_b2
    project = PEProject.objects.get(id=project_id)
    docs = project.documents.filter(local_file__isnull=False)
    
    count = 0
    for doc in docs:
        try:
            local_path = doc.local_file.path
            upload_local_file_to_b2(
                local_path=local_path,
                b2_key=doc.file_key,
                content_type=doc.mime_type
            )
            # Update doc to reflect it's now on B2 (clear local_file)
            doc.local_file = None
            doc.is_confirmed = True
            doc.save()
            count += 1
        except Exception as e:
            logger.error(f"Failed to move document {doc.id} to B2: {e}")
            
    return f"Moved {count} documents to B2 for project {project.legal_name}"


# ---------------------------------------------------------------------------
# Phase 2: AI Valuation, Term Sheet, SPA Draft Tasks
# ---------------------------------------------------------------------------

@shared_task
def generate_ai_valuation(project_id, user_id=None):
    """
    Uses Gemini to generate DCF and LBO assumptions based on project financials,
    then runs calculate_dcf() and calculate_lbo() with those assumptions.
    """
    from .valuation import calculate_dcf, calculate_lbo
    from django.contrib.auth import get_user_model
    User = get_user_model()

    def _safe_float(val, default=0.0):
        if isinstance(val, (int, float)): return float(val)
        if isinstance(val, str):
            try: return float(val.replace(',', '').replace('%', ''))
            except: return default
        if isinstance(val, dict):
            for k in ['value', 'total', 'amount', 'rate', 'target']:
                if k in val: return _safe_float(val[k], default)
        return default

    try:
        project = PEProject.objects.get(id=project_id)
        user = User.objects.get(id=user_id) if user_id else None
        client = AIModelClient()

        # 1. Update progress
        progress = project.analysis_progress or {}
        progress['Valuation'] = 'processing'
        project.analysis_progress = progress
        project.save()

        # 2. Gather context from existing analyses
        financials = list(project.financials.values(
            'fiscal_year_bs', 
            'revenue_npr', 
            'ebitda_npr', 
            'net_profit_npr',
            'total_assets_npr', 
            'total_debt_npr'
        ))
        qoe = project.qoe_reports.first()
        commercial = project.commercial_analyses.first()
        scoring = project.scoring_runs.first()

        context_data = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "deal_type": project.get_deal_type_display(),
            "financials": json.dumps(financials, default=str),
            "qoe_status": qoe.status if qoe else "N/A",
            "qoe_adjustments": json.dumps(qoe.analysis_data if qoe else {}, default=str),
            "commercial_notes": commercial.market_positioning_notes if commercial else "N/A",
            "scoring_total": str(scoring.total_deal_score) if scoring else "N/A",
            "nepal_context": "Nepal market. Currency: NPR. Risk-free rate ~8-10%. Country risk premium 4-6%.",
        }

        # 2. Call AI for assumption generation
        raw_json = client.execute_task("valuation_generation", context_data, project=project)
        
        # Parse JSON
        import re
        try:
            # Extract outermost JSON object to bypass any conversational text
            match = re.search(r'\{.*\}', raw_json, re.DOTALL)
            if match:
                clean_json = match.group(0)
            else:
                clean_json = raw_json
                
            # Clean markdown if regex failed to isolate cleanly (e.g., if there's text inside the matched braces... wait, re.search r'\{.*\}' grabs from first { to last })
            data = json.loads(clean_json)
        except Exception as e:
            logger.error(f"Valuation JSON Parse Error: {e}. Raw JSON: {raw_json}")
            data = {}

        def _get_any(d, keys, default=None):
            for k in keys:
                if k in d: return d[k]
            return default

        # 3. Process DCF Assumptions
        dcf_raw = data.get("dcf", {})
        dcf_assumptions = {
            "current_revenue": _safe_float(_get_any(dcf_raw, ["current_revenue", "revenue"]), 1000000),
            "revenue_growth_rate": _safe_float(_get_any(dcf_raw, ["revenue_growth_rate", "revenue_growth", "growth_rate"]), 0.15),
            "ebitda_margin": _safe_float(_get_any(dcf_raw, ["ebitda_margin", "margin"]), 0.20),
            "tax_rate": _safe_float(_get_any(dcf_raw, ["tax_rate", "tax"]), 0.25),
            "projection_years": int(_safe_float(_get_any(dcf_raw, ["projection_years", "years", "exit_year"]), 5)),
            "wacc": _safe_float(_get_any(dcf_raw, ["wacc", "discount_rate"]), 0.14),
            "terminal_growth_rate": _safe_float(_get_any(dcf_raw, ["terminal_growth_rate", "terminal_growth"]), 0.03),
            "net_debt": _safe_float(_get_any(dcf_raw, ["net_debt", "debt"]), 0),
        }

        dcf_outputs = calculate_dcf(dcf_assumptions)

        dcf_model = ValuationModel.objects.create(
            project=project,
            model_type='DCF',
            assumptions=dcf_assumptions,
            outputs=dcf_outputs,
            created_by=user
        )

        # 4. Process LBO Assumptions
        lbo_raw = data.get("lbo", {})
        lbo_assumptions = {
            "entry_revenue": _safe_float(_get_any(lbo_raw, ["entry_revenue", "revenue"]), dcf_assumptions["current_revenue"]),
            "entry_ebitda": _safe_float(_get_any(lbo_raw, ["entry_ebitda", "ebitda"]), dcf_assumptions["current_revenue"] * dcf_assumptions["ebitda_margin"]),
            "entry_multiple": _safe_float(_get_any(lbo_raw, ["entry_multiple", "multiple"]), 8.0),
            "exit_multiple": _safe_float(_get_any(lbo_raw, ["exit_multiple", "target_multiple"]), 10.0),
            "exit_year": int(_safe_float(_get_any(lbo_raw, ["exit_year", "years"]), 5)),
            "revenue_growth": _safe_float(_get_any(lbo_raw, ["revenue_growth", "growth"]), dcf_assumptions["revenue_growth_rate"]),
            "ebitda_margin": _safe_float(_get_any(lbo_raw, ["ebitda_margin", "margin"]), dcf_assumptions["ebitda_margin"]),
            "tax_rate": _safe_float(_get_any(lbo_raw, ["tax_rate", "tax"]), 0.25),
            "debt_financing": _get_any(lbo_raw, ["debt_financing", "debt"], [
                {"name": "Senior Debt", "amount": dcf_assumptions["net_debt"] or 1000000, "rate": 0.12}
            ])
        }
        lbo_outputs = calculate_lbo(lbo_assumptions)

        lbo_model = ValuationModel.objects.create(
            project=project,
            model_type='LBO',
            assumptions=lbo_assumptions,
            outputs=lbo_outputs,
            created_by=user
        )

        # 6. Store AI rationale
        ai_rationale = data.get("rationale", "AI-generated assumptions based on project financials and sector benchmarks.")

        # 7. Update progress
        progress = project.analysis_progress or {}
        progress['Valuation'] = 'completed'
        project.analysis_progress = progress
        project.save()

        _log_audit_event(
            ImmutableAuditEvent.EventType.VALUATION_OVERRIDE,
            project,
            actor=user,
            payload={
                'method': 'AI_GENERATED',
                'dcf_equity_value': dcf_outputs.get('equity_value'),
                'lbo_irr': lbo_outputs.get('irr'),
                'rationale': ai_rationale
            }
        )

        logger.info(f"AI Valuation complete for {project.legal_name}")

        return {
            "dcf_id": str(dcf_model.id),
            "lbo_id": str(lbo_model.id),
            "rationale": ai_rationale,
        }

    except Exception as e:
        logger.error(f"Error in generate_ai_valuation: {str(e)}")
        # Reset progress to prevent UI from sticking in loading state
        try:
            progress = project.analysis_progress or {}
            progress['Valuation'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e


@shared_task
def generate_ai_term_sheet(project_id, user_id=None):
    """
    Uses Gemini to generate an initial term sheet based on valuation models,
    scoring results, and deal context.
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()

    try:
        project = PEProject.objects.get(id=project_id)
        project.refresh_from_db() # Get latest valuation metrics
        
        user = User.objects.get(id=user_id) if user_id else None
        client = AIModelClient()

        # 1. Gather context
        dcf = project.valuations.filter(model_type='DCF').first()
        lbo = project.valuations.filter(model_type='LBO').first()
        scoring = project.scoring_runs.order_by('-created_at').first()
        fund = project.fund
        memo = project.memos.order_by('-version').first()

        # 1. Update Progress
        progress = project.analysis_progress or {}
        progress['Term Sheet'] = 'processing'
        project.analysis_progress = progress
        project.save(update_fields=['analysis_progress'])

        context_data = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "business_description": project.business_description,
            "investment_memo_summary": memo.content.get('executive_summary', 'N/A') if memo and memo.content else "N/A",
            "dcf_enterprise_value": str(dcf.outputs.get("enterprise_value", "N/A")) if dcf else "N/A",
            "dcf_equity_value": str(dcf.outputs.get("equity_value", "N/A")) if dcf else "N/A",
            "lbo_irr": str(lbo.outputs.get("irr", "N/A")) if lbo else "N/A",
            "lbo_moic": str(lbo.outputs.get("moic", "N/A")) if lbo else "N/A",
            "scoring_total": str(scoring.total_deal_score) if scoring else "N/A",
            "fund_name": fund.name,
            "fund_target_size": str(fund.target_size_npr),
            "investment_range_min": str(project.investment_range_min_npr or "N/A"),
            "investment_range_max": str(project.investment_range_max_npr or "N/A"),
            "nepal_context": "Nepal PE market. Currency: NPR. Standard PE terms for growth capital.",
        }
        
        logger.info(f"Generating AI Term Sheet for {project.legal_name} with context: {context_data}")

        # 2. Call AI
        raw_json = client.execute_task("term_sheet_draft", context_data, project=project)

        # 3. Parse
        clean = raw_json or ""
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            parts = clean.split("```")
            if len(parts) > 2:
                clean = parts[1].strip()
            
        try:
            if not clean:
                raise ValueError("AI returned an empty response.")
            data = json.loads(clean)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse term sheet JSON. Raw response: {raw_json[:500]}...")
            raise ValueError(f"AI returned malformed JSON. Raw: {raw_json[:500]}") from e

        # 4. Get version
        latest = project.term_sheets.order_by('-version').first()
        new_version = (latest.version + 1) if latest else 1

        # 5. Save
        term_sheet = TermSheet.objects.create(
            project=project,
            terms=data,
            ai_generated_terms=data,
            version=new_version,
            status='DRAFT',
            created_by=user,
        )

        logger.info(f"AI Term Sheet v{new_version} generated for {project.legal_name}")
        
        # Clear progress
        progress = project.analysis_progress or {}
        if 'Term Sheet' in progress:
            del progress['Term Sheet']
        project.analysis_progress = progress
        project.save(update_fields=['analysis_progress'])
        
        return {"term_sheet_id": str(term_sheet.id), "version": new_version}

    except Exception as e:
        logger.error(f"Error in generate_ai_term_sheet: {str(e)}")
        # Clear progress on error
        progress = project.analysis_progress or {}
        if 'Term Sheet' in progress:
            del progress['Term Sheet']
        project.analysis_progress = progress
        project.save(update_fields=['analysis_progress'])
        raise e


@shared_task
def generate_ai_spa_draft(project_id, user_id=None):
    """
    Uses Gemini to generate an initial SPA (Share Purchase Agreement) draft
    based on term sheet, regulatory context, and deal details.
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()

    try:
        project = PEProject.objects.get(id=project_id)
        user = User.objects.get(id=user_id) if user_id else None
        client = AIModelClient()

        # 1. Gather context
        term_sheet = project.term_sheets.order_by('-version').first()
        regulatory = getattr(project, 'regulatory_checklist', None)
        fund = project.fund

        context_data = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "deal_type": project.get_deal_type_display(),
            "term_sheet_terms": json.dumps(term_sheet.terms if term_sheet else {}, default=str),
            "fund_name": fund.name,
            "fitta_required": str(regulatory.fitta_approval_required) if regulatory else "Unknown",
            "nrb_required": str(regulatory.nrb_approval_required) if regulatory else "Unknown",
            "nepal_context": (
                "Nepal jurisdiction. Governing law: Nepal Contract Act 2056. "
                "Company Act 2063 applies. Securities Act 2063 for listed entities. "
                "Foreign Investment and Technology Transfer Act (FITTA) 2075 for FDI."
            ),
        }

        # 2. Call AI
        raw_json = client.execute_task("spa_draft", context_data, project=project)

        # 3. Parse
        clean = raw_json
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()

        data = json.loads(clean)



        # 4. Get version
        latest = project.spa_drafts.order_by('-version').first()
        new_version = (latest.version + 1) if latest else 1

        # 5. Save
        spa = SPADraft.objects.create(
            project=project,
            sections=data,
            ai_generated_sections=data,
            version=new_version,
            status='DRAFT',
            created_by=user,
        )

        logger.info(f"AI SPA Draft v{new_version} generated for {project.legal_name}")
        
        # Clear progress
        progress = project.analysis_progress or {}
        if 'SPA Draft' in progress:
            del progress['SPA Draft']
        project.analysis_progress = progress
        project.save(update_fields=['analysis_progress'])

        return {"spa_draft_id": str(spa.id), "version": new_version}

    except Exception as e:
        logger.error(f"Error in generate_ai_spa_draft: {str(e)}")
        # Clear progress on error
        progress = project.analysis_progress or {}
        if 'SPA Draft' in progress:
            del progress['SPA Draft']
        project.analysis_progress = progress
        project.save(update_fields=['analysis_progress'])
        raise e


import json
import logging
import time
import fitz  # PyMuPDF
import requests
from celery import shared_task, chain
import os
from django.conf import settings

from django.db import models
from django.utils import timezone
from deals.models import (
    PEProject, 
    PEProjectDocument, 
    ExtractedFinancials, 
    QoEReport,
    CommercialAnalysis,
    OperationalAnalysis,
    AICallLog,
    RedFlagPattern,
    RedFlagFinding,
    ScoringRun,
    CriterionScore,
    ComplianceGate,
    RegulatoryChecklist,
    DealMemo,
    Fund,
    LPFundCommitment,
    LPProfile,
    FundDocument,
    ValuationModel,
    TermSheet,
    SPADraft,
    ImmutableAuditEvent,
)
from deals.pdf_utils import generate_capital_account_pdf, upload_pdf_to_b2
from deals.mail_utils import send_statement_notification, send_capital_call_notification
from deals.signals import _log_audit_event



import re


from deals.ai_client import AIModelClient
from deals.b2_utils import generate_presigned_download_url

logger = logging.getLogger(__name__)

from .deal import generate_memo_draft

from deals.models import PromptLibrary


def extract_text_from_pdf(pdf_content):
    """Extracts text from PDF bytes using PyMuPDF."""
    text = ""
    try:
        with fitz.open(stream=pdf_content, filetype="pdf") as doc:
            for page in doc:
                text += page.get_text()
    except Exception as e:
        logger.error(f"PyMuPDF error: {e}")
    return text


def extract_text_from_excel(excel_content):
    """Extracts text from Excel bytes using openpyxl."""
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed, skipping Excel extraction")
        return ""
    
    text = ""
    try:
        wb = load_workbook(io.BytesIO(excel_content), data_only=True)
        total_chars = 0
        MAX_CHARS = 500000 # 500k char safety limit
        
        for sheet in wb.worksheets:
            if total_chars > MAX_CHARS:
                break
                
            text += f"\nSheet: {sheet.title}\n"
            # Limit to first 1000 rows per sheet to avoid huge calculation sheets
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i > 1000:
                    text += "... (Large sheet truncated)\n"
                    break
                
                # Filter out None values and clean up row
                row_data = [str(cell).strip() for cell in row if cell is not None]
                if not any(row_data): continue # Skip empty rows
                
                row_text = " | ".join(row_data)
                text += row_text + "\n"
                total_chars += len(row_text)
                    
                if total_chars > MAX_CHARS:
                    text += "\n... (Total document truncated for AI safety)\n"
                    break
    except Exception as e:
        logger.error(f"Excel extraction error: {e}")
    return text


def extract_text_from_pptx(pptx_content):
    """Extracts text from PowerPoint (.pptx) bytes using native zipfile and xml parsing."""
    import zipfile
    import io
    import xml.etree.ElementTree as ET
    
    text_runs = []
    try:
        z = zipfile.ZipFile(io.BytesIO(pptx_content))
        
        # Get all slide files in order
        slide_files = sorted(
            [f for f in z.namelist() if f.startswith("ppt/slides/slide") and f.endswith(".xml")],
            key=lambda x: int(''.join(filter(str.isdigit, x)))
        )
        
        namespace = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
        
        for slide_file in slide_files:
            slide_xml = z.read(slide_file)
            root = ET.fromstring(slide_xml)
            slide_text = []
            for text_node in root.findall('.//a:t', namespace):
                if text_node.text:
                    slide_text.append(text_node.text)
            if slide_text:
                slide_num = ''.join(filter(str.isdigit, slide_file))
                text_runs.append(f"--- Slide {slide_num} ---\n" + " ".join(slide_text))
                
        z.close()
    except Exception as e:
        logger.error(f"Native PPTX extraction error: {e}")
        return ""
    return "\n\n".join(text_runs)


def extract_text_from_docx(docx_content):
    """Extracts text from Word (.docx) bytes using native zipfile and xml parsing."""
    import zipfile
    import io
    import xml.etree.ElementTree as ET
    
    text_runs = []
    try:
        z = zipfile.ZipFile(io.BytesIO(docx_content))
        # Word document main body XML resides in word/document.xml
        if "word/document.xml" in z.namelist():
            doc_xml = z.read("word/document.xml")
            root = ET.fromstring(doc_xml)
            namespace = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            
            for text_node in root.findall('.//w:t', namespace):
                if text_node.text:
                    text_runs.append(text_node.text)
        z.close()
    except Exception as e:
        logger.error(f"Native DOCX extraction error: {e}")
        return ""
    return " ".join(text_runs)



@shared_task(name='deals.tasks.extract_financials_from_document')
def extract_financials_from_document(document_id):
    """
    Task to extract 3+ years of financial data from a PDF document.
    """
    try:
        doc = PEProjectDocument.objects.get(id=document_id)
        project = doc.project
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Extraction'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        logger.info(f"Starting financial extraction for document: {doc.filename}")
        
        # 1. Get document content
        file_content = None
        if doc.local_file:
            try:
                # In Docker, local_file points to /app/media/... which is volume mounted
                with doc.local_file.open('rb') as f:
                    file_content = f.read()
                logger.info(f"Read local file for {doc.filename}")
            except Exception as e:
                logger.warning(f"Could not read local file {doc.local_file}: {e}")
        
        if not file_content:
            logger.info(f"Downloading from B2: {doc.file_key}")
            download_url = generate_presigned_download_url(doc.file_key)
            response = requests.get(download_url)
            if response.status_code == 200:
                file_content = response.content
            else:
                raise Exception(f"Failed to download from B2. Status: {response.status_code}")
            
        # 2. Extract text based on file type
        file_ext = doc.filename.split('.')[-1].lower()
        if file_ext in ['xlsx', 'xlsm', 'xltx', 'xltm']:
            doc_text = extract_text_from_excel(file_content)
        elif file_ext in ['docx', 'docm', 'dotx', 'dotm']:
            doc_text = extract_text_from_docx(file_content)
        elif file_ext in ['pptx', 'pptm', 'potx', 'potm']:
            doc_text = extract_text_from_pptx(file_content)
        else:
            doc_text = extract_text_from_pdf(file_content)
            
        if not doc_text or len(doc_text) < 100:
            # Fallback for scanned PDFs or empty files
            logger.warning(f"No/Low text extracted from {doc.filename}. Relying on AI's native OCR.")
            doc_text = "[SCANNED DOCUMENT DETECTED - PLEASE ANALYZE ATTACHED FILE VISUALLY]"
        
        # 3. Call AI Client
        client = AIModelClient()
        context_data = {
            "document_text": doc_text,
            "project_name": project.legal_name
        }
        
        raw_json = client.execute_task("financial_extraction", context_data, project=project, document=doc)
        
        # 4. Parse and Save
        # Expecting JSON format: {"fiscal_years": [{"fiscal_year_bs": "2080/81", "revenue": 100, ...}], "confidence": 0.9}
        try:
            # Clean JSON if AI wrapped it in markdown code blocks
            if "```json" in raw_json:
                raw_json = raw_json.split("```json")[1].split("```")[0].strip()
            elif "```" in raw_json:
                raw_json = raw_json.split("```")[1].split("```")[0].strip()
                
            data = json.loads(raw_json)
        except Exception as e:
            logger.error(f"Failed to parse AI JSON output: {raw_json}")
            raise Exception(f"AI output parsing error: {str(e)}")

        fiscal_years = data.get("fiscal_years", [])
        if not fiscal_years:
            # Try to handle direct list if AI didn't nest it
            if isinstance(data, list):
                fiscal_years = data
            else:
                raise Exception("No fiscal years found in AI output")

        for entry in fiscal_years:
            ExtractedFinancials.objects.update_or_create(
                project=project,
                fiscal_year_bs=entry.get("fiscal_year_bs"),
                defaults={
                    "source_document": doc,
                    "revenue_npr": entry.get("revenue", 0),
                    "ebitda_npr": entry.get("ebitda", 0),
                    "net_profit_npr": entry.get("net_profit", entry.get("pat", 0)),
                    "total_assets_npr": entry.get("total_assets", 0),
                    "total_debt_npr": entry.get("total_debt", 0),
                    # Backwards compatibility
                    "revenue": entry.get("revenue", 0),
                    "ebitda": entry.get("ebitda", 0),
                    "pat": entry.get("net_profit", entry.get("pat", 0)),
                    "extraction_confidence": data.get("confidence", 0.85),
                    "raw_ai_output": entry
                }
            )
            
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Extraction'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Successfully extracted {len(fiscal_years)} years of financials for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Error in extract_financials_from_document: {str(e)}")
        # Mark as failed in progress
        try:
            progress = project.analysis_progress or {}
            progress['Extraction'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e


@shared_task(name='deals.tasks.run_qoe_analysis')
def run_qoe_analysis(project_id):
    """
    Task to run Quality of Earnings analysis using DeepSeek R1.
    """
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['QoE'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        financials = project.financials.all().order_by('fiscal_year_bs')
        
        if not financials.exists():
            raise Exception("No extracted financials found for QoE analysis")
            
        # 1. Prepare data for AI
        financial_summary = "Financial Data History:\n"
        for f in financials:
            financial_summary += (
                f"FY {f.fiscal_year_bs}: Revenue={f.revenue_npr}, EBITDA={f.ebitda_npr}, "
                f"PAT={f.net_profit_npr}, Assets={f.total_assets_npr}, Debt={f.total_debt_npr}\n"
            )
            
        # 2. Call AI Client (Routed to DeepSeek R1)
        client = AIModelClient()
        context_data = {
            "financial_data": financial_summary,
            "project_name": project.legal_name,
            "sector": project.get_sector_display()
        }
        
        report_text = client.execute_task("qoe_analysis", context_data, project=project)
        
        # 3. Determine Status (Red/Yellow/Green)
        # We can ask the AI for a structured status or parse it.
        # For now, let's use a robust keyword search or assume AI includes it.
        status = 'CLEAN'
        lower_report = report_text.lower()
        if any(word in lower_report for word in ['high risk', 'severe', 'red flag']):
            status = 'HIGH_RISK'
        elif any(word in lower_report for word in ['caution', 'yellow flag', 'moderate risk', 'adjustment needed']):
            status = 'CAUTION'
            
        # 4. Create Report
        QoEReport.objects.filter(project=project).delete()
        QoEReport.objects.create(
            project=project,
            report_text=report_text,
            status=status,
            analysis_data={
                "generated_at": str(timezone.now()),
                "fiscal_years_analyzed": [f.fiscal_year_bs for f in financials]
            }
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['QoE'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"QoE Report generated for {project.legal_name} (Status: {status})"
        
    except Exception as e:
        logger.error(f"Error in run_qoe_analysis: {str(e)}")
        try:
            progress = project.analysis_progress or {}
            progress['QoE'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e



@shared_task(name='deals.tasks.run_commercial_analysis')
def run_commercial_analysis(project_id):
    """
    Analyzes customer concentration and market positioning.
    """
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Prepare context
        client = AIModelClient()
        
        # Aggregate multiple form responses for a richer context
        all_responses = project.form_responses.all()
        business_desc = project.legal_name
        market_info = ""
        competition = ""
        usp = ""
        
        for r in all_responses:
            step = r.step_name.lower()
            data_str = str(r.response_data)
            if 'info' in step or 'overview' in step: business_desc = data_str
            if 'market' in step: market_info = data_str
            if 'competi' in step: competition = data_str
            if 'usp' in step or 'advantage' in step: usp = data_str

        # --- NEW: SCAN PITCH DECK/BUSINESS PLAN ---
        # Expanded to include COMMERCIAL category and filename heuristics
        pitch_deck = project.documents.filter(
            models.Q(is_confirmed=True) & (
                models.Q(category__in=['PITCH_DECK', 'BUSINESS_PLAN', 'COMMERCIAL']) |
                models.Q(filename__icontains='deck') |
                models.Q(filename__icontains='pitch')
            )
        ).first()
        
        pitch_deck_content = ""
        if pitch_deck:
            logger.info(f"--- ATTEMPTING PITCH DECK SCAN: {pitch_deck.filename} ---")
            try:
                # 1. Get file content
                deck_content = None
                if pitch_deck.local_file:
                    try:
                        with pitch_deck.local_file.open('rb') as f:
                            deck_content = f.read()
                    except Exception as e:
                        logger.warning(f"Could not read local pitch deck: {e}")
                
                if not deck_content:
                    download_url = generate_presigned_download_url(pitch_deck.file_key)
                    response = requests.get(download_url)
                    if response.status_code == 200:
                        deck_content = response.content
                
                # 2. Extract text based on file extension
                deck_text = ""
                if deck_content:
                    deck_ext = pitch_deck.filename.split('.')[-1].lower()
                    if deck_ext in ['xlsx', 'xlsm']:
                        deck_text = extract_text_from_excel(deck_content)
                    elif deck_ext == 'pdf':
                        deck_text = extract_text_from_pdf(deck_content)
                    elif deck_ext in ['pptx', 'pptm', 'potx', 'potm', 'ppt']:
                        deck_text = extract_text_from_pptx(deck_content)
                    elif deck_ext in ['docx', 'docm', 'dotx', 'dotm', 'doc']:
                        deck_text = extract_text_from_docx(deck_content)
                
                if not deck_text or len(deck_text) < 50:
                    deck_text = "[No readable text found in document slides]"
                
                # We reuse the AI client to summarize the pitch deck specifically for commercial strategy, passing actual extracted text
                pitch_deck_content = client.execute_task(
                    "document_qualitative_summary", 
                    {
                        "filename": pitch_deck.filename,
                        "document_text": deck_text
                    }, 
                    project=project,
                    document=pitch_deck
                )
                logger.info(f"--- PITCH DECK SCAN SUCCESS: {len(pitch_deck_content)} chars ---")
            except Exception as scan_err:
                logger.error(f"--- PITCH DECK SCAN FAILED: {str(scan_err)} ---")
                pitch_deck_content = "Document scanning failed. Relying on form responses."
        else:
            logger.warning("--- NO PITCH DECK FOUND FOR SCANNING ---")

        context_data = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "business_desc": business_desc,
            "market_info": market_info,
            "competition": competition,
            "usp": usp,
            "pitch_deck_insights": pitch_deck_content,
            "market_data": f"Sector: {project.get_sector_display()}. Analysis requested for Nepal market context. {market_info} {pitch_deck_content}",
        }

        # Call AI (Routed to Gemini)
        raw_json = client.execute_task("commercial_analysis", context_data, project=project)
        
        try:
            if "```json" in raw_json:
                raw_json = raw_json.split("```json")[1].split("```")[0].strip()
            data = json.loads(raw_json)
        except:
            data = {"customer_concentration_pct": 0, "top_customer_names": "", "market_positioning_notes": raw_json}

        # Create Record
        CommercialAnalysis.objects.filter(project=project).delete()
        CommercialAnalysis.objects.create(
            project=project,
            customer_concentration_pct=data.get("customer_concentration_pct", 0),
            top_customer_names=data.get("top_customer_names", ""),
            market_positioning_notes=data.get("market_positioning_notes", ""),
            ai_call_log=AICallLog.objects.filter(project=project, task_type='commercial_analysis').first()
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Commercial'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Commercial Analysis completed for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Error in run_commercial_analysis: {str(e)}")
        try:
            progress = project.analysis_progress or {}
            progress['Commercial'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e



@shared_task(name='deals.tasks.run_operational_analysis')
def run_operational_analysis(project_id, manual_context=""):
    """
    Analyzes technology stack and operational risks.
    """
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Operational'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        client = AIModelClient()
        
        # 1. Aggregate Operational Context from Form Responses
        tech_info = ""
        team_info = ""
        ops_info = ""
        
        responses = project.form_responses.all()
        for resp in responses:
            step = resp.step_name.lower()
            data_str = json.dumps(resp.response_data)
            if 'tech' in step or 'product' in step: tech_info = data_str
            if 'team' in step or 'founder' in step: team_info = data_str
            if 'operation' in step or 'supply' in step: ops_info = data_str

        # --- NEW: SCAN DOCUMENTS FOR OPERATIONAL DATA ---
        op_doc = project.documents.filter(
            models.Q(is_confirmed=True) & (
                models.Q(category__in=['PITCH_DECK', 'BUSINESS_PLAN', 'TECHNICAL', 'COMMERCIAL', 'OPERATIONAL_AUDIT', 'TECH_STACK', 'ORG_CHART']) |
                models.Q(filename__icontains='tech') |
                models.Q(filename__icontains='ops')
            )
        ).first()
        
        doc_insights = ""
        if op_doc:
            try:
                # 1. Get file content
                doc_content = None
                if op_doc.local_file:
                    try:
                        with op_doc.local_file.open('rb') as f:
                            doc_content = f.read()
                    except Exception as e:
                        logger.warning(f"Could not read local operational doc: {e}")
                
                if not doc_content:
                    download_url = generate_presigned_download_url(op_doc.file_key)
                    response = requests.get(download_url)
                    if response.status_code == 200:
                        doc_content = response.content
                
                # 2. Extract text based on file extension
                doc_text = ""
                if doc_content:
                    doc_ext = op_doc.filename.split('.')[-1].lower()
                    if doc_ext in ['xlsx', 'xlsm']:
                        doc_text = extract_text_from_excel(doc_content)
                    elif doc_ext == 'pdf':
                        doc_text = extract_text_from_pdf(doc_content)
                    elif doc_ext in ['pptx', 'pptm', 'potx', 'potm', 'ppt']:
                        doc_text = extract_text_from_pptx(doc_content)
                    elif doc_ext in ['docx', 'docm', 'dotx', 'dotm', 'doc']:
                        doc_text = extract_text_from_docx(doc_content)
                
                if not doc_text or len(doc_text) < 50:
                    doc_text = "[No readable text found in document]"
                
                doc_insights = client.execute_task(
                    "document_qualitative_summary", 
                    {
                        "filename": op_doc.filename,
                        "document_text": doc_text,
                        "focus": "technology stack, operational logic, and founder expertise"
                    }, 
                    project=project,
                    document=op_doc
                )
            except Exception as scan_err:
                logger.error(f"--- OPERATIONAL SCAN FAILED: {str(scan_err)} ---")
                doc_insights = "No document insights available. Document scanning failed."

        context_data = {
            "project_name": project.legal_name,
            "tech_stack_raw": tech_info,
            "team_details": team_info,
            "operational_notes": ops_info,
            "manual_context": manual_context,
            "document_insights": doc_insights,
            "market_sector": project.get_sector_display()
        }

        # 2. Call AI (Generates JSON + Markdown)
        # We append a strict formatting instruction to ensure the AI always provides the JSON block
        context_data["formatting_instruction"] = """
        IMPORTANT: Your response MUST end with a JSON block enclosed in ```json ... ``` tags.
        The JSON must follow this exact structure:
        {
          "technology_stack": { "Frontend": "...", "Backend": "...", "Database": "...", "Infrastructure": "..." },
          "key_person_risk_score": 1-10 integer,
          "supply_chain_risks": ["Risk 1", "Risk 2"],
          "operational_red_flags": ["Blocker 1"]
        }
        Do not include any text after the JSON block.
        """
        
        raw_output = client.execute_task("operational_analysis", context_data, project=project)
        
        data = {
            "technology_stack": {}, 
            "key_person_risk_score": 5, 
            "supply_chain_risks": [], 
            "operational_red_flags": [],
            "thesis_markdown": raw_output
        }

        try:
            # 1. Try standard markdown code block extraction
            json_str = None
            if "```json" in raw_output:
                json_str = raw_output.split("```json")[1].split("```")[0].strip()
            elif "{" in raw_output and "}" in raw_output:
                # Fallback: Find the last JSON-like structure in the text
                import re
                matches = re.findall(r'\{.*\}', raw_output, re.DOTALL)
                if matches:
                    json_str = matches[-1]

            if json_str:
                parsed = json.loads(json_str)
                data.update(parsed)
                # Aggressively clean up markdown: remove everything from the first ```json to the end
                if "```json" in raw_output:
                    data["thesis_markdown"] = raw_output.split("```json")[0].strip()
                elif json_str in raw_output:
                    data["thesis_markdown"] = raw_output.split(json_str)[0].strip()
            else:
                data["operational_red_flags"] = ["AI provided insights but skipped structured metrics. Please review thesis for details."]
        except Exception as e:
            logger.warning(f"Operational JSON parsing failed: {e}")

        # Enforce defaults and fallback values if parsed data is invalid or None
        tech_stack = data.get("technology_stack", {})
        if not isinstance(tech_stack, dict):
            tech_stack = {}
            
        key_score = data.get("key_person_risk_score", 5)
        if key_score is None:
            key_score = 5
        else:
            try:
                key_score = int(float(key_score))
            except:
                key_score = 5

        supply_chain = data.get("supply_chain_risks", [])
        if not isinstance(supply_chain, list):
            supply_chain = []
            
        red_flags = data.get("operational_red_flags", [])
        if not isinstance(red_flags, list):
            red_flags = []

        # Create Record
        OperationalAnalysis.objects.filter(project=project).delete()
        OperationalAnalysis.objects.create(
            project=project,
            technology_stack=tech_stack,
            key_person_risk_score=key_score,
            supply_chain_risks=supply_chain,
            operational_red_flags=red_flags,
            thesis_markdown=data.get("thesis_markdown", raw_output),
            ai_call_log=AICallLog.objects.filter(project=project, task_type='operational_analysis').first()
        )
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Operational'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"Operational Analysis completed for {project.legal_name}"
        
    except Exception as e:
        logger.error(f"Error in run_operational_analysis: {str(e)}")
        try:
            progress = project.analysis_progress or {}
            progress['Operational'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e


@shared_task(name='deals.tasks.scan_legal_document')
def scan_legal_document(document_id):
    """
    Scans a legal document for red flags using regex patterns + Gemini Flash analysis.
    """
    try:
        doc_obj = PEProjectDocument.objects.get(id=document_id)
        project = doc_obj.project
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Legal Scan'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        # 1. Download and Extract Text
        url = generate_presigned_download_url(doc_obj.file_key)
        resp = requests.get(url)
        if resp.status_code != 200:
            raise Exception(f"Failed to download document: {resp.status_code}")
            
        file_ext = doc_obj.filename.split('.')[-1].lower()
        if file_ext in ['docx', 'docm', 'dotx', 'dotm', 'doc']:
            text = extract_text_from_docx(resp.content)
        elif file_ext in ['xlsx', 'xlsm', 'xltx', 'xltm']:
            text = extract_text_from_excel(resp.content)
        elif file_ext in ['pptx', 'pptm', 'potx', 'potm', 'ppt']:
            text = extract_text_from_pptx(resp.content)
        else:
            text = extract_text_from_pdf(resp.content)
        if not text:
            return "No text extracted from document."

        # 2. Regex Matching
        patterns = RedFlagPattern.objects.all()
        findings_to_create = []
        
        client = AIModelClient()
        
        for p in patterns:
            matches = list(re.finditer(p.pattern_regex, text, re.IGNORECASE))
            if matches:
                # Take first match context for AI analysis
                match = matches[0]
                start = max(0, match.start() - 200)
                end = min(len(text), match.end() + 200)
                snippet = text[start:end]
                
                # Use Gemini to analyze the specific snippet
                ai_prompt = {
                    "snippet": snippet,
                    "pattern_name": p.name,
                    "description": p.description,
                    "nepal_context": p.nepal_context_note
                }
                
                analysis = client.execute_task("legal_scan", ai_prompt, project=project)
                
                findings_to_create.append(RedFlagFinding(
                    project=project,
                    document=doc_obj,
                    pattern=p,
                    context_snippet=snippet,
                    severity=p.severity,
                    ai_analysis=analysis
                ))

        if findings_to_create:
            RedFlagFinding.objects.bulk_create(findings_to_create)
            
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Legal Scan'] = 'completed'
        project.analysis_progress = progress
        project.save()
            
        return f"Legal scan completed for {doc_obj.filename}. Found {len(findings_to_create)} red flags."

    except Exception as e:
        logger.error(f"Error in scan_legal_document: {str(e)}")
        try:
            progress = project.analysis_progress or {}
            progress['Legal Scan'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e



@shared_task(name='deals.tasks.run_finlo_scoring')
def run_finlo_scoring(project_id, user_id=None):
    """
    Executes the 20-criteria FINLO scoring framework.
    Runs 5 parallel AI calls (one per pillar).
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    try:
        project = PEProject.objects.get(id=project_id)
        
        # Update progress
        progress = project.analysis_progress or {}
        progress['Scoring'] = 'processing'
        project.analysis_progress = progress
        project.save()
        
        user = User.objects.get(id=user_id) if user_id else None
        
        # 1. Create Scoring Run
        run = ScoringRun.objects.create(
            project=project,
            triggered_by=user,
            status='PROCESSING'
        )
        
        # 2. Gather Context
        financials = list(project.financials.all())
        qoe = project.qoe_reports.first()
        red_flags = list(project.red_flags.all())
        responses = list(project.form_responses.all())
        operational = project.operational_analyses.first()
        commercial = project.commercial_analyses.first()
        
        context = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "deal_type": project.get_deal_type_display(),
            "financials": [
                {
                    "year": f.fiscal_year_bs,
                    "rev": float(f.revenue),
                    "ebitda": float(f.ebitda),
                    "pat": float(f.pat)
                } for f in financials
            ],
            "qoe_status": qoe.status if qoe else "N/A",
            "legal_red_flags": [f"{f.pattern.name}: {f.severity}" for f in red_flags if f.pattern],
            "operational_thesis": operational.thesis_markdown if operational else "No operational audit performed yet.",
            "operational_red_flags": operational.operational_red_flags if operational else [],
            "commercial_notes": commercial.market_positioning_notes if commercial else "No commercial audit performed yet.",
            "submission_data": [str(r.response_data) for r in responses]
        }
        
        # 3. Define Pillars
        pillars = {
            "F": {
                "name": "Foresight",
                "criteria": ["problem_clarity", "market_opportunity", "innovation_type", "ai_native_assessment"]
            },
            "I": {
                "name": "Insight",
                "criteria": ["market_size_accuracy", "revenue_model_clarity", "unit_economics_awareness", "capital_efficiency"]
            },
            "N": {
                "name": "Nexus",
                "criteria": ["founder_market_fit", "team_completeness", "leadership_conviction", "network_strength"]
            },
            "L": {
                "name": "Logic",
                "criteria": ["validation_evidence", "traction_quality", "risk_awareness", "comparable_analysis"]
            },
            "O": {
                "name": "Odyssey",
                "criteria": ["partnership_quality", "competitive_moat", "growth_vision", "finlogic_alignment"]
            }
        }
        
        client = AIModelClient()
        pillar_averages = {}
        
        # 4. Execute Consolidated AI Call
        # We consolidate all 5 pillars into 1 call for 70% token savings and holistic reasoning.
        prompt_data = {
            "pillars_definition": pillars,
            "project_data": context,
            "precision_guardrails": (
                "1. Analyze all 20 criteria across the 5 pillars (F, I, N, L, O).\n"
                "2. Every score MUST be an integer between 1-10.\n"
                "3. Provide specific evidence quotes for EACH criterion.\n"
                "4. Ensure cross-pillar consistency: if a Red Flag exists in Legal, it must reflect in Risk Awareness scoring.\n"
                "5. BREVITY: Keep rationales to max 2 concise, high-impact sentences. Avoid filler text."
            )
        }
        
        # This uses the 'scoring' task type in PromptLibrary
        raw_json = client.execute_task("scoring", prompt_data, project=project)
        
        try:
            # Aggressive markdown cleanup
            clean_json = raw_json
            if "```json" in clean_json:
                clean_json = clean_json.split("```json")[1].split("```")[0].strip()
            elif "```" in clean_json:
                clean_json = clean_json.split("```")[1].split("```")[0].strip()
            
            data = json.loads(clean_json)
            pillar_averages = {}
            
            # The AI might return the results nested under 'pillar_results' or at the root
            pillar_results = data.get("pillar_results", data)
            
            for code, info in pillars.items():
                # Try to find pillar data by code (F, I, N, L, O) or by name (Foresight, etc.)
                pillar_data = pillar_results.get(code, pillar_results.get(info["name"], {}))
                
                # If the AI returned criteria_scores at the root of the pillar object
                criteria_data = pillar_data.get("criteria_scores", pillar_data)
                
                scores_sum = 0
                count = 0
                
                for crit_id in info["criteria"]:
                    result = criteria_data.get(crit_id, {})
                    score_val = result.get("score", 5)
                    
                    CriterionScore.objects.create(
                        scoring_run=run,
                        pillar=code,
                        criterion_id=crit_id,
                        ai_score=score_val,
                        ai_rationale=result.get("rationale", "No rationale provided."),
                        ai_confidence=result.get("confidence", 0.7),
                        evidence_quotes=result.get("evidence", [])
                    )
                    scores_sum += score_val
                    count += 1
                
                pillar_averages[code] = (scores_sum / count) if count > 0 else 5

        except Exception as e:
            logger.error(f"Failed to parse consolidated scoring response: {str(e)}")
            
            # CRITICAL DEBUG: Write the failing response to a file so we can analyze it
            try:
                debug_path = os.path.join(settings.BASE_DIR, "scratch", "last_failed_scoring.json")
                os.makedirs(os.path.dirname(debug_path), exist_ok=True)
                with open(debug_path, "w", encoding="utf-8") as f:
                    f.write(raw_json)
                logger.info(f"Failing AI response dumped to: {debug_path}")
            except:
                pass

            # Create default entries so the UI isn't empty
            for code, info in pillars.items():
                for crit_id in info["criteria"]:
                    CriterionScore.objects.create(
                        scoring_run=run,
                        pillar=code,
                        criterion_id=crit_id,
                        ai_score=5,
                        ai_rationale="Consolidated scoring failed to parse. Using baseline defaults.",
                        ai_confidence=0.1
                    )
                pillar_averages[code] = 5

        # 5. Calculate Weighted Final Score
        # F: 20%, I: 25%, N: 20%, L: 20%, O: 15%
        final_score = (
            pillar_averages.get("F", 5) * 0.20 +
            pillar_averages.get("I", 5) * 0.25 +
            pillar_averages.get("N", 5) * 0.20 +
            pillar_averages.get("L", 5) * 0.20 +
            pillar_averages.get("O", 5) * 0.15
        )
        
        run.total_deal_score = round(final_score, 2)
        run.status = 'COMPLETED'
        run.save()
        
        # 6. Initialize Compliance Gates
        gates = ['FITTA', 'AML_KYC', 'FINANCIAL_AUDIT', 'LEGAL_STRUCTURE', 'SEBON_MAPPING']
        for g in gates:
            ComplianceGate.objects.get_or_create(scoring_run=run, gate_id=g)
            
        # Note: Scoring no longer changes project status automatically.
        # GP advances to IC_REVIEW manually after reviewing scoring + clearing compliance gates.
        project.save()
        
        # Finalize progress
        progress = project.analysis_progress or {}
        progress['Scoring'] = 'completed'
        project.analysis_progress = progress
        project.save()
        
        return f"FINLO Scoring completed for {project.legal_name}. Score: {run.total_deal_score}"
        
    except Exception as e:
        logger.error(f"Error in run_finlo_scoring: {str(e)}")
        try:
            progress = project.analysis_progress or {}
            progress['Scoring'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        if 'run' in locals():
            run.status = 'FAILED'
            run.save()
        raise e



@shared_task(name='deals.tasks.run_nepal_compliance_check')
def run_nepal_compliance_check(project_id):
    """
    Analyzes project context to pre-populate the Nepal Regulatory Checklist.
    Detects FITTA, NRB, and SEBON requirements.
    """
    project = PEProject.objects.get(id=project_id)
    
    # 1. Detect Foreign Investment (FITTA)
    # Checks: "foreign" in company name, or if there's a specific form response
    is_foreign = False
    if "foreign" in project.legal_name.lower():
        is_foreign = True
        
    # Check form responses for keywords
    for resp in project.form_responses.all():
        data_str = json.dumps(resp.response_data).lower()
        if any(kw in data_str for kw in ["fdi", "foreign", "non-nepali", "international investor"]):
            is_foreign = True
            break
            
    # 2. Detect NRB Approval (usually for financial/banking/large scale FDI)
    is_nrb = is_foreign # Most FDI needs NRB recording/approval
    
    # 3. Create or Update Checklist
    checklist, created = RegulatoryChecklist.objects.get_or_create(
        project=project,
        defaults={
            "fitta_approval_required": is_foreign,
            "nrb_approval_required": is_nrb,
            "sebon_reporting_compliant": True, # Assume compliant initially
            "industry_specific_license_required": project.sector in [PEProject.Sector.MANUFACTURING, PEProject.Sector.AGRICULTURE]
        }
    )
    
    if not created:
        # Update if not manually locked (could add a lock flag later)
        checklist.fitta_approval_required = is_foreign
        checklist.nrb_approval_required = is_nrb
        checklist.save()
        
    return f"Compliance checklist pre-populated for {project.legal_name}"



@shared_task(name='deals.tasks.run_full_analysis')
def run_full_analysis(project_id, user_id=None):
    """
    Triggers the complete AI due diligence pipeline in sequence.
    """
    try:
        from deals.models import PEProjectDocument
        project = PEProject.objects.get(id=project_id)
        
        # Initialize Progress
        project.analysis_progress = {
            "Extraction": "pending",
            "QoE": "pending",
            "Commercial": "pending",
            "Operational": "pending",
            "Compliance": "pending",
            "Legal Scan": "pending",
            "Scoring": "pending",
            "Memo": "pending"
        }
        project.save()
        
        # 1. Find the most recent financial document
        doc = project.documents.filter(is_confirmed=True, category__in=['FINANCIAL', 'FINANCIALS']).order_by('-uploaded_at').first()
        legal_doc = project.documents.filter(is_confirmed=True, category='LEGAL').order_by('-uploaded_at').first()
        
        if not doc:
            # Let's set progress to failed so user is notified why it aborted
            progress = project.analysis_progress or {}
            for k in progress:
                progress[k] = 'failed'
            project.analysis_progress = progress
            project.save()
            return f"Full analysis aborted: No financial document found for {project.legal_name}"

        # 2. Build the chain dynamically
        steps = [
            extract_financials_from_document.si(doc.id),
            run_qoe_analysis.si(project.id),
            run_commercial_analysis.si(project.id),
            run_operational_analysis.si(project.id),
            run_nepal_compliance_check.si(project.id)
        ]
        
        if legal_doc:
            steps.append(scan_legal_document.si(legal_doc.id))
            
        steps.append(run_finlo_scoring.si(project.id))
        steps.append(generate_memo_draft.si(project.id))
        
        analysis_chain = chain(*steps)
        
        analysis_chain.apply_async()
        
        # Log Audit Event
        from deals.signals import _log_audit_event
        from django.contrib.auth import get_user_model
        User = get_user_model()
        actor = User.objects.filter(id=user_id).first() if user_id else None

        _log_audit_event(
            event_type='FULL_ANALYSIS_TRIGGERED',
            actor=actor,
            obj=project,
            payload={"document_id": str(doc.id) if doc else None}
        )
        
        return f"Full analysis chain started for {project.legal_name}"
    except Exception as e:
        logger.error(f"Error in run_full_analysis: {str(e)}")
        # Mark all steps as failed
        try:
            from deals.models import PEProject
            project = PEProject.objects.get(id=project_id)
            progress = project.analysis_progress or {}
            for k in progress:
                if progress[k] in ['processing', 'pending']:
                    progress[k] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e



@shared_task(name='deals.tasks.generate_ai_valuation')
def generate_ai_valuation(project_id, user_id=None):
    """
    Uses Gemini to generate DCF and LBO assumptions based on project financials,
    then runs calculate_dcf() and calculate_lbo() with those assumptions.
    """
    from deals.valuation import calculate_dcf, calculate_lbo
    from django.contrib.auth import get_user_model
    User = get_user_model()

    def _safe_float(val, default=0.0):
        if isinstance(val, (int, float)): return float(val)
        if isinstance(val, str):
            try: return float(val.replace(',', '').replace('%', ''))
            except: return default
        if isinstance(val, dict):
            for k in ['value', 'total', 'amount', 'rate', 'target']:
                if k in val: return _safe_float(val[k], default)
        return default

    try:
        project = PEProject.objects.get(id=project_id)
        user = User.objects.get(id=user_id) if user_id else None
        client = AIModelClient()

        # 1. Update progress
        progress = project.analysis_progress or {}
        progress['Valuation'] = 'processing'
        project.analysis_progress = progress
        project.save()

        # 2. Gather context from existing analyses
        financials = list(project.financials.values(
            'fiscal_year_bs', 
            'revenue_npr', 
            'ebitda_npr', 
            'net_profit_npr',
            'total_assets_npr', 
            'total_debt_npr'
        ))
        qoe = project.qoe_reports.first()
        commercial = project.commercial_analyses.first()
        scoring = project.scoring_runs.first()

        context_data = {
            "project_name": project.legal_name,
            "sector": project.get_sector_display(),
            "deal_type": project.get_deal_type_display(),
            "financials": json.dumps(financials, default=str),
            "qoe_status": qoe.status if qoe else "N/A",
            "qoe_adjustments": json.dumps(qoe.analysis_data if qoe else {}, default=str),
            "commercial_notes": commercial.market_positioning_notes if commercial else "N/A",
            "scoring_total": str(scoring.total_deal_score) if scoring else "N/A",
            "nepal_context": "Nepal market. Currency: NPR. Risk-free rate ~8-10%. Country risk premium 4-6%.",
        }

        # 2. Call AI for assumption generation
        raw_json = client.execute_task("valuation_generation", context_data, project=project)
        
        # Parse JSON
        import re
        try:
            # Extract outermost JSON object to bypass any conversational text
            match = re.search(r'\{.*\}', raw_json, re.DOTALL)
            if match:
                clean_json = match.group(0)
            else:
                clean_json = raw_json
                
            # Clean markdown if regex failed to isolate cleanly (e.g., if there's text inside the matched braces... wait, re.search r'\{.*\}' grabs from first { to last })
            data = json.loads(clean_json)
        except Exception as e:
            logger.error(f"Valuation JSON Parse Error: {e}. Raw JSON: {raw_json}")
            data = {}

        def _get_any(d, keys, default=None):
            for k in keys:
                if k in d: return d[k]
            return default

        # 3. Process DCF Assumptions
        dcf_raw = data.get("dcf", {})
        dcf_assumptions = {
            "current_revenue": _safe_float(_get_any(dcf_raw, ["current_revenue", "revenue"]), 1000000),
            "revenue_growth_rate": _safe_float(_get_any(dcf_raw, ["revenue_growth_rate", "revenue_growth", "growth_rate"]), 0.15),
            "ebitda_margin": _safe_float(_get_any(dcf_raw, ["ebitda_margin", "margin"]), 0.20),
            "tax_rate": _safe_float(_get_any(dcf_raw, ["tax_rate", "tax"]), 0.25),
            "projection_years": int(_safe_float(_get_any(dcf_raw, ["projection_years", "years", "exit_year"]), 5)),
            "wacc": _safe_float(_get_any(dcf_raw, ["wacc", "discount_rate"]), 0.14),
            "terminal_growth_rate": _safe_float(_get_any(dcf_raw, ["terminal_growth_rate", "terminal_growth"]), 0.03),
            "net_debt": _safe_float(_get_any(dcf_raw, ["net_debt", "debt"]), 0),
        }

        dcf_outputs = calculate_dcf(dcf_assumptions)

        dcf_model = ValuationModel.objects.create(
            project=project,
            model_type='DCF',
            assumptions=dcf_assumptions,
            outputs=dcf_outputs,
            created_by=user
        )

        # 4. Process LBO Assumptions
        lbo_raw = data.get("lbo", {})
        lbo_assumptions = {
            "entry_revenue": _safe_float(_get_any(lbo_raw, ["entry_revenue", "revenue"]), dcf_assumptions["current_revenue"]),
            "entry_ebitda": _safe_float(_get_any(lbo_raw, ["entry_ebitda", "ebitda"]), dcf_assumptions["current_revenue"] * dcf_assumptions["ebitda_margin"]),
            "entry_multiple": _safe_float(_get_any(lbo_raw, ["entry_multiple", "multiple"]), 8.0),
            "exit_multiple": _safe_float(_get_any(lbo_raw, ["exit_multiple", "target_multiple"]), 10.0),
            "exit_year": int(_safe_float(_get_any(lbo_raw, ["exit_year", "years"]), 5)),
            "revenue_growth": _safe_float(_get_any(lbo_raw, ["revenue_growth", "growth"]), dcf_assumptions["revenue_growth_rate"]),
            "ebitda_margin": _safe_float(_get_any(lbo_raw, ["ebitda_margin", "margin"]), dcf_assumptions["ebitda_margin"]),
            "tax_rate": _safe_float(_get_any(lbo_raw, ["tax_rate", "tax"]), 0.25),
            "debt_financing": _get_any(lbo_raw, ["debt_financing", "debt"], [
                {"name": "Senior Debt", "amount": dcf_assumptions["net_debt"] or 1000000, "rate": 0.12}
            ])
        }
        lbo_outputs = calculate_lbo(lbo_assumptions)

        lbo_model = ValuationModel.objects.create(
            project=project,
            model_type='LBO',
            assumptions=lbo_assumptions,
            outputs=lbo_outputs,
            created_by=user
        )

        # 6. Store AI rationale
        ai_rationale = data.get("rationale", "AI-generated assumptions based on project financials and sector benchmarks.")

        # 7. Update progress
        progress = project.analysis_progress or {}
        progress['Valuation'] = 'completed'
        project.analysis_progress = progress
        project.save()

        _log_audit_event(
            ImmutableAuditEvent.EventType.VALUATION_OVERRIDE,
            project,
            actor=user,
            payload={
                'method': 'AI_GENERATED',
                'dcf_equity_value': dcf_outputs.get('equity_value'),
                'lbo_irr': lbo_outputs.get('irr'),
                'rationale': ai_rationale
            }
        )

        logger.info(f"AI Valuation complete for {project.legal_name}")

        return {
            "dcf_id": str(dcf_model.id),
            "lbo_id": str(lbo_model.id),
            "rationale": ai_rationale,
        }

    except Exception as e:
        logger.error(f"Error in generate_ai_valuation: {str(e)}")
        # Reset progress to prevent UI from sticking in loading state
        try:
            progress = project.analysis_progress or {}
            progress['Valuation'] = 'failed'
            project.analysis_progress = progress
            project.save()
        except:
            pass
        raise e


